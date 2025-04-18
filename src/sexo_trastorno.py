from collections import defaultdict

from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from abrir_principal import cargar_principal
from form_tabla import crear_tabla


def reporte_distribucion_sexo_trastornos(filename):
    wb, ws = cargar_principal(filename)
    if not wb or not ws:
        print("Error: No se pudo cargar el archivo principal.")
        return

    ws_reporte = wb.create_sheet("Distribucion_Sexos_Trastornos")

    # Definir sexos de dato
    sexos_registrados = {
        "Masculino": 0,
        "Femenino": 0,
        "No binario": 0,
    }

    conteo_trastornos_por_sexo = defaultdict(lambda: defaultdict(int))

    # Estoy empezando a considerar pedir los trastornos como parametro, pero eso podria enredar el programa, lo dejare asi de momento
    # Obtener los encabezados de trastornos
    encabezados_trastornos = [
        ws.cell(row=1, column=col).value
        for col in range(7, ws.max_column + 1)
        if ws.cell(row=1, column=col).value
    ]

    for row in range(2, ws.max_row + 1):
        # Un try si se fue un dato que no sea int a la casilla
        try:
            dato = ws.cell(row=row, column=3).value
        except (TypeError, ValueError):
            continue

        if dato == "Masculino":
            sexo = "Masculino"

        elif dato == "Femenino":
            sexo = "Femenino"

        else:
            sexo = "No binario"

        sexos_registrados[sexo] += 1

        for col, trastorno in enumerate(encabezados_trastornos, start=7):
            if (
                ws.cell(row=row, column=col).value
                and ws.cell(row=row, column=col).value.strip() == "Si"
            ):
                conteo_trastornos_por_sexo[sexo][trastorno] += 1

    #################### Tabla y grafico de distribucion de datoes  ####################
    # Encabezados
    ws_reporte.append(["Sexo", "Cantidad de Personas"] + encabezados_trastornos)

    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    # Aplicar estilos a los encabezados
    for celda in ws_reporte[1]:
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = thin_border

    # Agregar datos de distribución de datoes y trastornos
    for row_num, (sexo, cantidad) in enumerate(sexos_registrados.items(), 2):
        fila = [sexo, cantidad] + [
            conteo_trastornos_por_sexo[sexo][t] for t in encabezados_trastornos
        ]
        ws_reporte.append(fila)

        for col_num, dato in enumerate(fila, 1):
            celda = ws_reporte.cell(row=row_num, column=col_num, value=dato)
            celda.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            celda.border = thin_border

    # Ajustar ancho de columnas
    for col in ws_reporte.iter_cols():
        col_letter = get_column_letter(col[0].column)
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws_reporte.column_dimensions[col_letter].width = min(max_length + 2, 30)

    # Gráfico de Barras
    bar_chart = BarChart()
    bar_chart.title = "Distribución de Sexos"
    bar_chart.x_axis.title = "Sexos"
    bar_chart.y_axis.title = "Cantidad"
    bar_chart.width = 22
    bar_chart.height = 12

    data = Reference(
        ws_reporte, min_col=2, min_row=1, max_row=len(sexos_registrados) + 1
    )
    categories = Reference(
        ws_reporte, min_col=1, min_row=2, max_row=len(sexos_registrados) + 1
    )

    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    ws_reporte.add_chart(bar_chart, "B12")

    # Gráfico de Pastel
    pie_chart = PieChart()
    pie_chart.title = "Distribución de Sexos"
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(categories)
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True
    pie_chart.width = 22
    pie_chart.height = 12
    ws_reporte.add_chart(pie_chart, "B36")

    #################### Tabla y grafico de promedio de aflicciones por edad  ####################

    # Cálculo de promedios
    ws_reporte.append([""] * (len(encabezados_trastornos) + 2))

    # Encabezados
    ws_reporte.append(
        ["Sexo", "Cantidad de Personas"]
        + [f"Promedio {t}" for t in encabezados_trastornos]
    )

    start_row_promedios = (
        len(sexos_registrados) + 3
    )  # Fila donde inician los datos de la segunda tabla

    # Aplicar estilos a los encabezados, se aplican los mismos estilos que en la primera tabla
    for col in range(1, len(encabezados_trastornos) + 3):
        celda = ws_reporte.cell(row=start_row_promedios, column=col)
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = thin_border

    # Llenar la tabla con los promedios y aplicar estilos basicos
    for row_num, (rango, cantidad) in enumerate(
        sexos_registrados.items(), start_row_promedios + 1
    ):
        if cantidad > 0:
            fila_promedios = [rango, cantidad] + [
                round(conteo_trastornos_por_sexo[rango][t] / cantidad, 2)
                for t in encabezados_trastornos
            ]
        else:
            fila_promedios = [rango, cantidad] + [0] * len(encabezados_trastornos)

        ws_reporte.append(fila_promedios)

        for col_num, dato in enumerate(fila_promedios, 1):
            celda = ws_reporte.cell(row=row_num, column=col_num, value=dato)
            celda.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            celda.border = thin_border

    crear_tabla(
        ws_reporte,
        type_table="TableStyleMedium2",
        table_name="TablaPromediosTrastornosPorSexo",
    )

    # Gráfico de Barras Apiladas - Promedio de Trastornos por Sexo
    bar_chart = BarChart()
    bar_chart.title = "Promedio de trastornos por Sexoes"
    bar_chart.y_axis.title = "Promedio de personas que padecen"
    bar_chart.x_axis.title = "Sexo"
    bar_chart.height = 30
    bar_chart.width = 40

    data = Reference(
        ws_reporte,
        min_col=3,
        min_row=len(sexos_registrados) + 3,
        max_row=len(sexos_registrados) + 9,
        max_col=len(encabezados_trastornos) + 2,
    )

    categories = Reference(
        ws_reporte,
        min_col=1,
        min_row=len(sexos_registrados) + 4,
        max_row=len(sexos_registrados) + 9,
    )

    # Agregar los datos al grafico de barras
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)

    # Agregar el grafico a la hoja del reporte
    ws_reporte.add_chart(bar_chart, "F12")

    from openpyxl.chart import RadarChart

    # Gráfico de Radar - Promedio de Trastornos por Sexo
    radar_chart = RadarChart()
    radar_chart.title = "Distribución de Promedio de Trastornos por Sexo"
    radar_chart.style = 13
    radar_chart.y_axis.title = "Promedio"
    radar_chart.x_axis.title = "Trastornos"
    radar_chart.height = 30
    radar_chart.width = 40

    # Seleccionar los datos
    data = Reference(
        ws_reporte,
        min_col=3,
        min_row=len(sexos_registrados) + 3,
        max_row=len(sexos_registrados) + 6,
        max_col=len(encabezados_trastornos) + 2,
    )

    categories = Reference(
        ws_reporte,
        min_col=1,
        min_row=len(sexos_registrados) + 4,
        max_row=len(sexos_registrados) + 6,
    )

    radar_chart.add_data(data, titles_from_data=True)
    radar_chart.set_categories(categories)

    # Agregar la gráfica de radar al reporte
    ws_reporte.add_chart(radar_chart, "O12")

    # Guardar archivo
    output_path = filename
    try:
        wb.save(output_path)
        wb.close()
        print(f"Distribucion sexo - trastorno guardado en: '{output_path}'")
    except Exception as e:
        print(f"Error al guardar el archivo: {e}")


if __name__ == "__main__":
    reporte_distribucion_sexo_trastornos()
