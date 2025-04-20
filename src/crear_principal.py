from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from form_tabla import crear_tabla


# Función para calcular la edad de las personas con su fecha de nacimiento
def calcular_edad(fecha_nacimiento):
    if not fecha_nacimiento or type(fecha_nacimiento) == str:
        return ""
    hoy = datetime.today()

    # check if fecha_nacimiento is exactly Jan, 1, 1900, if so, return empty string
    if (
        fecha_nacimiento.year == 1900
        and fecha_nacimiento.month == 1
        and fecha_nacimiento.day == 1
    ):
        return ""

    edad = (
        hoy.year
        - fecha_nacimiento.year
        - ((hoy.month, hoy.day) < (fecha_nacimiento.month, fecha_nacimiento.day))
    )

    return edad


from typing import Dict

# 1. Diccionario de mapeo: claves en minúsculas (abreviaciones y nombres completos)
_STATE_MAP: Dict[str, str] = {
    # Aguascalientes
    "ags": "Aguascalientes",
    "agua": "Aguascalientes",
    "aguascalientes": "Aguascalientes",
    # Baja California
    "bc": "Baja California",
    "baja california": "Baja California",
    "bcn": "Baja California",
    # Baja California Sur
    "bcs": "Baja California Sur",
    "baja california sur": "Baja California Sur",
    # Campeche
    "camp": "Campeche",
    "campeche": "Campeche",
    # Coahuila
    "coah": "Coahuila de Zaragoza",
    "coahuila": "Coahuila de Zaragoza",
    # Colima
    "col": "Colima",
    "colima": "Colima",
    # Chiapas
    "chis": "Chiapas",
    "chiapas": "Chiapas",
    # Chihuahua
    "chih": "Chihuahua",
    "chihuahua": "Chihuahua",
    # Ciudad de México
    "cdmx": "Ciudad de México",
    "ciudad de méxico": "Ciudad de México",
    # Durango
    "dur": "Durango",
    "durango": "Durango",
    # Guanajuato
    "gto": "Guanajuato",
    "guanajuato": "Guanajuato",
    # Guerrero
    "gro": "Guerrero",
    "guerrero": "Guerrero",
    # Hidalgo
    "hgo": "Hidalgo",
    "hidalgo": "Hidalgo",
    # Jalisco
    "jal": "Jalisco",
    "jalisco": "Jalisco",
    # México (Estado de México)
    "em": "Estado de México",
    "mex": "Estado de México",
    "estado de méxico": "Estado de México",
    # Michoacán
    "mich": "Michoacán",
    "michoacán": "Michoacán",
    # Morelos
    "mor": "Morelos",
    "morelos": "Morelos",
    # Nayarit
    "nay": "Nayarit",
    "nayarit": "Nayarit",
    # Nuevo León
    "nl": "Nuevo León",
    "nuevo león": "Nuevo León",
    # Oaxaca
    "oax": "Oaxaca",
    "oaxaca": "Oaxaca",
    # Puebla
    "pue": "Puebla",
    "puebla": "Puebla",
    # Querétaro
    "qro": "Querétaro",
    "querétaro": "Querétaro",
    # Quintana Roo
    "qroo": "Quintana Roo",
    "quintana roo": "Quintana Roo",
    # San Luis Potosí
    "slp": "San Luis Potosí",
    "san luis potosí": "San Luis Potosí",
    # Sinaloa
    "sin": "Sinaloa",
    "sinaloa": "Sinaloa",
    # Sonora
    "son": "Sonora",
    "sonora": "Sonora",
    # Tabasco
    "tab": "Tabasco",
    "tabasco": "Tabasco",
    # Tamaulipas
    "tamps": "Tamaulipas",
    "tamaulipas": "Tamaulipas",
    # Tlaxcala
    "tlax": "Tlaxcala",
    "tlaxcala": "Tlaxcala",
    # Veracruz
    "ver": "Veracruz de Ignacio de la Llave",
    "veracruz": "Veracruz de Ignacio de la Llave",
    # Yucatán
    "yuc": "Yucatán",
    "yucatán": "Yucatán",
    # Zacatecas
    "zac": "Zacatecas",
    "zacatecas": "Zacatecas",
}


def normalize_state(input_str: str) -> str:
    """
    Convierte una cadena que representa un estado de México
    a su forma canónica con mayúscula inicial y acentos correctos.

    Parámetros:
        input_str (str): Abreviatura o nombre completo en cualquier formato de mayúsculas/minúsculas.

    Retorna:
        str: Nombre del estado normalizado. Si no coincide con ninguna clave del mapeo,
            se devuelve una cadena vacía.
    """
    key = input_str.strip().lower()
    return _STATE_MAP.get(
        key, "Sin datos"
    )  # Devuelve "Sin datos" si no se encuentra el estado


# Crear la tabla principal, donde se muestran los resultados de la encuesta y sus datos relevantes.
# No se incluyeron las respuestas de cada pregunta.
# La tabla generado aqui se utilizara para generar el resto de reportes, para no tener que estar
# llamando a la base de datos en cada reporte.
def crear_tabla_principal(data, filename):
    # Obtener los datos de cadda relevantes
    lineas = []
    for persona in data:
        fecha_nacimiento = persona.get("birthdate")
        try:
            edad = calcular_edad(fecha_nacimiento) if fecha_nacimiento else ""
        except ValueError:
            edad = ""
        # Normalizar estado de origen y residencia
        estado_origen = normalize_state(persona.get("stateOrigin", "") or "-")
        estado_residencia = normalize_state(persona.get("stateResidence", "") or "-")
        fila = [
            persona.get("name", ""),
            edad,
            persona.get("gender", ""),
            persona.get("sexualPreference", ""),
            estado_origen,
            estado_residencia,
            persona.get("diagnosticA1", ""),
            persona.get("diagnosticA2", ""),
            persona.get("diagnosticA3", ""),
            persona.get("diagnosticB1", ""),
            persona.get("diagnosticC1", ""),
            persona.get("riesgoC1", ""),
            persona.get("diagnosticD1", ""),
            persona.get("periodoD1", ""),
            persona.get("diagnosticD2", ""),
            persona.get("periodoD2", ""),
            persona.get("diagnosticE1", ""),
            persona.get("periodoE1", ""),
            persona.get("diagnosticE2", ""),
            persona.get("periodoE2", ""),
            persona.get("diagnosticF1", ""),
            persona.get("diagnosticF2", ""),
            persona.get("diagnosticF3", ""),
            persona.get("diagnosticG1", ""),
            persona.get("diagnosticH1", ""),
            persona.get("diagnosticI1", ""),
            persona.get("diagnosticJ1", ""),
            persona.get("diagnosticJ2", ""),
            persona.get("diagnosticK2", ""),
            persona.get("diagnosticK3", ""),
            persona.get("diagnosticL1", ""),
            persona.get("diagnosticL2", ""),
            persona.get("diagnosticL3", ""),
            persona.get("diagnosticM1", ""),
            persona.get("diagnosticN1", ""),
            persona.get("diagnosticN2", ""),
            persona.get("diagnosticO1", ""),
            persona.get("diagnosticP1", ""),
        ]
        lineas.append(fila)

    # Crear un nuevo libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Encuesta"

    # Definir encabezados de la tabla
    encabezados = [
        "Nombre",
        "Edad",
        "Sexo",
        "Preferencia sexual",
        "Estado de origen",
        "Estado de residencia",
    ]

    trastornos = [
        "Episodio depresivo mayor actual",
        "Episodio depresivo mayor recidivante",
        "Episodio depresivo mayor con síntomas melancólicos actual",
        "Trastorno distímico actual",
        "Riesgo de suicidio",
        "Riesgo",
        "Episodio hipomaníaco",
        "Periodo de episodio hipomaníaco",
        "Episodio maníaco",
        "Periodo de episodio maníaco",
        "Trastorno de angustia de por vida",
        "Periodo de trastorno de angustia",
        "Crisis actual con síntomas limitados",
        "Periodo de crisis",
        "Trastorno de angustia actual",
        "Trastorno de angustia sin agorafobia actual",
        "Trastorno de angustia con agorafobia actual",
        "Agorafobia actual sin historial de trastorno de angustia",
        "Fobia social actual",
        "Estado por estrés postraumático actual",
        "Dependencia de alcohol actual",
        "Abuso de alcohol actual",
        "Dependencia de sustancias actual",
        "Abuso de sustancias actual",
        "Trastorno psicótico actual",
        "Trastorno psicótico de por vida",
        "Trastorno del estado de ánimo con síntomas psicóticos actual",
        "Anorexia nerviosa actual",
        "Bulimia nerviosa actual",
        "Anorexia nerviosa tipo compulsivo/purgativo actual",
        "Trastorno de ansiedad generalizada actual",
        "Trastorno antisocial de la personalidad de por vida",
    ]

    encabezados += trastornos

    # Índices de los trastornos
    indice_trastornos = list(range(7, len(encabezados) + 1))

    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    # Agregar encabezados
    for col_num, encabezado in enumerate(encabezados, 1):
        celda = ws.cell(row=1, column=col_num, value=encabezado)
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = thin_border

    # Agregar los datos a la tabla
    for row_num, fila in enumerate(lineas, 2):
        if len(fila) != len(encabezados):
            fila += [""] * (
                len(encabezados) - len(fila)
            )  # En caso de no haber dato se rellena con un vacio

        # Contar cantidad de trastornos del entrevistado actual
        cantidad_si = sum(
            1
            for i in indice_trastornos
            if i - 1 < len(fila) and fila[i - 1].strip().lower() == "si"
        )
        # Definir color en función para darle al usuario en funcion de la cantidad de trastornos que padece
        if cantidad_si > 0:
            rojo = 255
            naranja = max(
                0, 255 - (cantidad_si * 25)
            )  # Disminuye progresivamente de 255 a 0

            color_hex = f"{rojo:02X}{naranja:02X}00"
            fill_color = PatternFill(start_color=color_hex, fill_type="solid")
        else:
            # No se pone color a usuarios que no sufren ningun trastorno
            fill_color = None

        for col_num, dato in enumerate(fila, 1):
            celda = ws.cell(row=row_num, column=col_num, value=dato)
            celda.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            celda.border = thin_border
            # Aplicar color de acuerdo a la cantidad de trastornos al nombre del entrevistado
            if fill_color and col_num == 1:
                celda.fill = fill_color
            # Se pintan de amarillos los "Si", para identificar más facilmente los trastornos padecidos
            if dato == "Si":
                celda.fill = PatternFill(start_color="FFD700", fill_type="solid")
            if dato == "Bajo":
                celda.fill = PatternFill(start_color="FFD700", fill_type="solid")
            if dato == "Moderado":
                celda.fill = PatternFill(start_color="FF8000", fill_type="solid")
            if dato == "Alto":
                celda.fill = PatternFill(start_color="FF0000", fill_type="solid")
            if dato == "actual":
                celda.fill = PatternFill(start_color="FF0000", fill_type="solid")
            if dato == "pasado":
                celda.fill = PatternFill(start_color="409ad6", fill_type="solid")

    # Ajustar el ancho de las columnas
    for col in range(1, len(encabezados) + 1):
        max_length = max(
            len(str(ws.cell(row=row, column=col).value or ""))
            for row in range(1, len(lineas) + 2)
        )
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(
            max_length + 2, 30
        )

    # Ajustar altura de filas
    for row in range(1, len(lineas) + 2):
        ws.row_dimensions[row].height = 40 if row == 1 else 20

    # Aplicar formato de tabla
    crear_tabla(ws, type_table="TableStyleLight11", table_name="Tabla_general")
    # Guardar el archivo Excel
    output_path = filename
    try:
        wb.save(output_path)
        wb.close()
        print(f"Página principal guardada en: '{output_path}'")
    except Exception as e:
        print(f"Error al guardar el archivo: {e}")


if __name__ == "__main__":
    crear_tabla_principal()
