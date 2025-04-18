from openpyxl.utils import get_column_letter
from abrir_principal import cargar_principal
from form_tabla import crear_tabla
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList

def reporte_caracteristicas_asociadas():
    wb, ws = cargar_principal()
    if not wb:
        return

    ws_reporte = wb.create_sheet("Factores_en_trastornos")

    # Indices de los datos de las columnas
    caracteristicas = {
        "Edad": 2,
        "Sexo": 3,
        "Preferencia sexual": 4,
        "Estado de origen": 5,
        "Estado de residencia": 6,
    }

    # Diccionario para contar cuántas veces aparece cada valor de cada característica entre personas con al menos un trastorno
    conteo_factores = {k: defaultdict(int) for k in caracteristicas}

    for row in range(2, ws.max_row + 1):
        tiene_trastorno = False
        for col in range(7, ws.max_column + 1):
            if ws.cell(row=row, column=col).value == "Si":
                tiene_trastorno = True
                break
        
        if tiene_trastorno:
            for nombre, col in caracteristicas.items():
                valor = ws.cell(row=row, column=col).value
                if valor:
                    conteo_factores[nombre][valor] += 1

    # Si no se encontraron factores asociados, eliminar la hoja y finalizar
    if not any(conteo_factores.values()):  # Verifica si todas las características están vacías
        wb.remove(ws_reporte)
        print("No se encontraron factores asociados a trastornos. La hoja ha sido eliminada.")
        return

    # Encabezados
    ws_reporte.append(["Característica", "Valor", "Cantidad de Personas con Trastornos"])

    for caracteristica, valores in conteo_factores.items():
        for valor, cantidad in valores.items():
            ws_reporte.append([caracteristica, valor, cantidad])

    # Estilos para los encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="medium"), right=Side(style="medium"),
                         top=Side(style="medium"), bottom=Side(style="medium"))

    # Aplicar estilos a los encabezados
    for celda in ws_reporte[1]:
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = thin_border

    for row in ws_reporte.iter_rows(min_row=2):
        for celda in row:
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.border = thin_border

    # Ajustar ancho de columnas
    for col in ws_reporte.iter_cols():
        col_letter = get_column_letter(col[0].column)
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws_reporte.column_dimensions[col_letter].width = min(max_length + 2, 30)
    
    crear_tabla(ws_reporte, type_table="TableStyleMedium2", table_name="TablaFactores")

    # Crear gráfico con los 10 valores más frecuentes en personas con trastornos
    datos_grafica = sorted(
        [(f"{c}: {v}", conteo) for c, valores in conteo_factores.items() for v, conteo in valores.items()],
        key=lambda x: x[1],
        reverse=True
    )[:10]

    if datos_grafica:
        start_row = len(ws_reporte["A"]) + 2
        ws_reporte.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
        celda_titulo = ws_reporte.cell(row=start_row, column=1, value="Top 10 Factores más Frecuentes en trastornos")
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
        celda_titulo.font = Font(bold=True, size=14)
        celda_titulo.fill = PatternFill(start_color="4F81BD", fill_type="solid", fgColor="4F81BD")
        ws_reporte.append(["Factor", "Cantidad de personas con trastornos"])
        for nombre, cantidad in datos_grafica:
            ws_reporte.append([nombre, cantidad])

        # Aplicar estilos al encabezado del top 10
        top10_header_row = start_row + 1
        for celda in ws_reporte[top10_header_row]:
            celda.font = header_font
            celda.fill = header_fill
            celda.alignment = header_alignment
            celda.border = thin_border

        # Aplicar estilos al contenido del top 10
        for row in ws_reporte.iter_rows(min_row=top10_header_row + 1, max_row=top10_header_row + len(datos_grafica)):
            for celda in row:
                celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                celda.border = thin_border

        # Grafico de barras para el top 10
        bar_chart = BarChart()
        bar_chart.title = "Factores Asociados a Trastornos"
        bar_chart.x_axis.title = "Factores"
        bar_chart.y_axis.title = "Frecuencia"
        bar_chart.height = 10
        bar_chart.width = 30

        data = Reference(ws_reporte, min_col=2, min_row=start_row + 1, max_row=start_row + len(datos_grafica) + 1)
        categories = Reference(ws_reporte, min_col=1, min_row=start_row + 2, max_row=start_row + len(datos_grafica) + 1)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(categories)

        ws_reporte.add_chart(bar_chart, f"F1")

        # Grafico de pastel para el top 10
        pie_chart = PieChart()
        pie_chart.title = "Distribución de Factores Asociados a Trastornos"

        # Referencias para los datos y categorías 
        data_pie = Reference(ws_reporte, min_col=2, min_row=start_row + 1, max_row=start_row + len(datos_grafica) + 1)
        categories_pie = Reference(ws_reporte, min_col=1, min_row=start_row + 2, max_row=start_row + len(datos_grafica) + 1)

        # Agregar datos al gráfico
        pie_chart.add_data(data_pie, titles_from_data=True)
        pie_chart.set_categories(categories_pie)

        # Etiquetas de porcentaje
        pie_chart.dataLabels = DataLabelList()
        pie_chart.dataLabels.showPercent = True  

        # Tamaño opcional 
        pie_chart.height = 15
        pie_chart.width = 20

        ws_reporte.add_chart(pie_chart, "F21")

    # Guardar los cambios en el archivo Excel
    output_path = "reporte_principal.xlsx"
    try:
        wb.save(output_path)
        wb.close()
        print(f"Página caracteristicas más comunes guardada en: '{output_path}'")
    except Exception as e:
        print(f"Error al guardar el archivo: {e}")

if __name__ == "__main__":
    reporte_caracteristicas_asociadas()
