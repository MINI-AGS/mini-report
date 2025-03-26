from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

def crear_tabla(ws, type_table, table_name):
    # ws es la hoja en la que se esta trabajando
    # Definir el rango de la tabla en base al tamaño de los datos
    max_col = ws.cell(row=1, column=ws.max_column).column_letter
    rango_tabla = f"A1:{max_col}{ws.max_row}" 

    # Crear la tabla
    tabla = Table(displayName=table_name, ref=rango_tabla)

    # Definir el estilo de la tabla
    estilo = TableStyleInfo(
        name=type_table,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )

    tabla.tableStyleInfo = estilo
    ws.add_table(tabla)