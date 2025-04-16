from openpyxl.utils import get_column_letter
from abrir_principal import cargar_principal
from form_tabla import crear_tabla
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

# Se genera una tabla y graficas de la distribucion del total de trastornos obtenidos
def reporte_distribucion_trastornos():
    wb, ws = cargar_principal()
    if not wb:
        return

    # Crear hoja para el reporte de distribución de trastornos
    ws_reporte = wb.create_sheet("Distribucion_trastornos")

    # Obtener los encabezados y los datos
    encabezados = [ws.cell(row=1, column=col).value for col in range(7, ws.max_column + 1) if ws.cell(row=1, column=col).value]  
    conteo_trastornos = defaultdict(int)

    # Contar la cantidad de veces que cada trastorno aparece con "Si"
    for row in range(2, ws.max_row + 1):
        for col, trastorno in enumerate(encabezados, start=7):  
            if ws.cell(row=row, column=col).value == "Si":
                conteo_trastornos[trastorno] += 1

    # Escribir la tabla en la hoja de reporte
    ws_reporte.append(["Trastorno", "Cantidad"])
    for trastorno, cantidad in conteo_trastornos.items():
        ws_reporte.append([trastorno, cantidad])

    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Estilos para bordes
    thin_border = Border(left=Side(style="medium"), right=Side(style="medium"),
                         top=Side(style="medium"), bottom=Side(style="medium"))

    # Aplicar estilos a los encabezados
    for celda in ws_reporte[1]:
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = thin_border

    # Escribir la tabla en la hoja de reporte 
    for row_num, (trastorno, cantidad) in enumerate(conteo_trastornos.items(), 2):
        fila = [trastorno, cantidad]
        
        for col_num, dato in enumerate(fila, 1):
            celda = ws_reporte.cell(row=row_num, column=col_num, value=dato)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.border = thin_border

    # Ajustar el ancho de las columnas 
    for col in ws_reporte.iter_cols():
        col_letter = get_column_letter(col[0].column)  
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws_reporte.column_dimensions[col_letter].width = min(max_length + 2, 30)

    # Dar formato de tabla
    crear_tabla(ws_reporte, type_table="TableStyleMedium2", table_name="TablaTrastornos")

    # Crear gráfico de barras
    bar_chart = BarChart()
    bar_chart.title = "Distribución de Trastornos"
    bar_chart.x_axis.title = "Trastornos"
    bar_chart.y_axis.title = "Cantidad"
    bar_chart.legend = None
    bar_chart.height = 10
    bar_chart.width = 30

    data = Reference(ws_reporte, min_col=2, min_row=1, max_row=len(conteo_trastornos) + 1)
    categories = Reference(ws_reporte, min_col=1, min_row=2, max_row=len(conteo_trastornos) + 1)
    
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    # Celda donde comienza la grafica (D2)
    ws_reporte.add_chart(bar_chart, "D2")  

    # Crear gráfico de pastel
    pie_chart = PieChart()
    pie_chart.title = "Distribución de Trastornos"
    pie_chart.height = 10
    pie_chart.width = 30

    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(categories)

    # Etiquetas de porcentaje
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True  

    # Celda donde comienza la grafica (D18)
    ws_reporte.add_chart(pie_chart, "D18")  

    # Guardar los cambios en el archivo Excel
    output_path = "reporte_principal.xlsx"
    try:
        wb.save(output_path)
        wb.close()
        print(f"Reporte distribucion - trastornos guardada en: '{output_path}'")
    except Exception as e:
        print(f"Error al guardar el archivo: {e}")

if __name__ == "__main__":
    reporte_distribucion_trastornos()