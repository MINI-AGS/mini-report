from openpyxl.utils import get_column_letter
from abrir_principal import cargar_principal  
from form_tabla import crear_tabla  
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList

def reporte_distribucion_trastornos_pref_sexual():
    wb, ws = cargar_principal()
    if not wb or not ws:
        print("Error: No se pudo cargar el archivo principal.")
        return

    ws_reporte = wb.create_sheet("Distribucion_pref_sexual")

    # Diccionario para contar las preferencias sexuales y trastornos
    pref_sexuals_registradas = defaultdict(int)
    conteo_trastornos_por_pref_sexual = defaultdict(lambda: defaultdict(int))

    # Obtener los encabezados de trastornos
    encabezados_trastornos = [ws.cell(row=1, column=col).value for col in range(7, ws.max_column + 1) if ws.cell(row=1, column=col).value]

    # Recorrer las filas para extraer datos
    for row in range(2, ws.max_row + 1):
        try:
            pref_sexual = ws.cell(row=row, column=4).value  # pref_sexual de origen (Columna 4)
        except (TypeError, ValueError):
            continue

        pref_sexuals_registradas[pref_sexual] += 1

        for col, trastorno in enumerate(encabezados_trastornos, start=7):
            if ws.cell(row=row, column=col).value and ws.cell(row=row, column=col).value.strip() == "Si":
                conteo_trastornos_por_pref_sexual[pref_sexual][trastorno] += 1

    #################### Creación de tabla y gráficos ####################

    # Encabezados
    ws_reporte.append(["pref_sexual", "Cantidad de Personas"] + encabezados_trastornos)

    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style="medium"), right=Side(style="medium"),
                         top=Side(style="medium"), bottom=Side(style="medium"))

    # Aplicar estilos a los encabezados
    for celda in ws_reporte[1]:
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = thin_border

    # Agregar datos de distribución
    for row_num, (pref_sexual, cantidad) in enumerate(pref_sexuals_registradas.items(), 2):
        fila = [pref_sexual, cantidad] + [conteo_trastornos_por_pref_sexual[pref_sexual][t] for t in encabezados_trastornos]
        ws_reporte.append(fila)

        for col_num, dato in enumerate(fila, 1):
            celda = ws_reporte.cell(row=row_num, column=col_num, value=dato)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.border = thin_border

    # Ajustar ancho de columnas
    for col in ws_reporte.iter_cols():
        col_letter = get_column_letter(col[0].column)
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws_reporte.column_dimensions[col_letter].width = min(max_length + 2, 30)

    # Gráfico de Barras
    bar_chart = BarChart()
    bar_chart.title = "Distribución de Personas por preferencia sexual"
    bar_chart.x_axis.title = "preferencia sexual"
    bar_chart.y_axis.title = "Cantidad de Personas"
    bar_chart.width = 22
    bar_chart.height = 12

    data = Reference(ws_reporte, min_col=2, min_row=1, max_row=len(pref_sexuals_registradas) + 1)
    categories = Reference(ws_reporte, min_col=1, min_row=2, max_row=len(pref_sexuals_registradas) + 1)

    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)
    fila_inicio = 2*(len(pref_sexuals_registradas) + 3) 
    ws_reporte.add_chart(bar_chart, f"B{fila_inicio}")

    # Gráfico de Pastel
    pie_chart = PieChart()
    pie_chart.title = "Distribución de Personas por preferencia sexual"
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(categories)
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True
    pie_chart.width = 22
    pie_chart.height = 12
    ws_reporte.add_chart(pie_chart, f"B{fila_inicio + 23}")

    #################### Tabla y grafico de promedio de aflicciones por edad  #################### 

    # Calculo de promedios
    ws_reporte.append([""] * (len(encabezados_trastornos) + 2))

    # Encabezados
    ws_reporte.append(["preferencia sexual", "Cantidad de Personas"] + [f"Promedio {t}" for t in encabezados_trastornos])

    # Aplicar estilos a los encabezados
    for col in range(1, len(encabezados_trastornos) + 3):
        celda = ws_reporte.cell(row=ws_reporte.max_row, column=col)
        celda.font = header_font    
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = thin_border
    
    # Llenar la tabla con los promedios y aplicar estilos basicos
    for row_num, (pref_sexual, cantidad) in enumerate(pref_sexuals_registradas.items(), ws_reporte.max_row + 1):
        ws_reporte.append([pref_sexual, cantidad] + [conteo_trastornos_por_pref_sexual[pref_sexual][t] / cantidad if cantidad != 0 else 0 for t in encabezados_trastornos])
        for col in range(1, len(encabezados_trastornos) + 3):
            celda = ws_reporte.cell(row=row_num, column=col)
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.border = thin_border

    # Aplicar formato de tabla
    crear_tabla(ws_reporte, type_table="TableStyleMedium2", table_name="Tabla_pref_sexual_origen")

    # Gráfico de Barras Apiladas - Promedio de Trastornos por Sexo
    bar_chart = BarChart()
    bar_chart.title = "Promedio de trastornos por preferencia sexual"
    bar_chart.y_axis.title = "Promedio de personas que padecen"
    bar_chart.x_axis.title = "preferencia sexual"
    bar_chart.height = 25
    bar_chart.width = 50

    # Seleccionar los datos
    data = Reference(
        ws_reporte,
        min_col=3,
        min_row=len(pref_sexuals_registradas) + 3,
        max_row=len(pref_sexuals_registradas) + len(pref_sexuals_registradas) + 3,
        max_col=len(encabezados_trastornos) + 2
    )

    categories = Reference(
        ws_reporte,
        min_col=1,
        min_row=len(pref_sexuals_registradas) + 4,
        max_row=len(pref_sexuals_registradas) + len(pref_sexuals_registradas) + 3
    )

    # Agregar los datos al grafico de barras
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(categories)

    # Agregar el grafico a la hoja del reporte
    ws_reporte.add_chart(bar_chart, f"O{fila_inicio}")

    # Crear gráfico de líneas en lugar de radar
    line_chart = LineChart()
    line_chart.title = "Distribución de Promedio de Trastornos por preferencia sexual"
    line_chart.style = 13
    line_chart.y_axis.title = "Promedio"
    line_chart.x_axis.title = "Trastornos"
    line_chart.height = 25  
    line_chart.width = 40

    # Seleccionar los datos
    data = Reference(
        ws_reporte,
        min_col=3,
        min_row=len(pref_sexuals_registradas) + 3,
        max_row=len(pref_sexuals_registradas) + len(pref_sexuals_registradas) + 3,
        max_col=len(encabezados_trastornos) + 2
    )

    categories = Reference(
        ws_reporte,
        min_col=1,
        min_row=len(pref_sexuals_registradas) + 4,
        max_row=len(pref_sexuals_registradas) + len(pref_sexuals_registradas) + 3
    )

    line_chart.add_data(data, titles_from_data=True)
    line_chart.set_categories(categories)
    line_chart.legend.position = "t"  # Posición de la leyenda: top

    # Agregar la gráfica de líneas al reporte
    ws_reporte.add_chart(line_chart, f"F{fila_inicio}")


    # Guardar el archivo Excel
    output_path = "reporte_principal.xlsx"
    try:
        wb.save(output_path)
        wb.close()
        print(f"Reporte pref_sexual - trstorno guardado en: '{output_path}'")
    except Exception as e:
        print(f"Error al guardar el archivo: {e}")

if __name__ == "__main__":
    reporte_distribucion_trastornos_pref_sexual()
